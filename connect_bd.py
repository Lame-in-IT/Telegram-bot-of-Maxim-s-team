import psycopg2
from config import config_wb as settings
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy_utils import database_exists, create_database
import pandas as pd
import logging
import json
import openpyxl
from openpyxl.styles.alignment import Alignment

# from get_date import get_date_7
# from api_shops import api_shops, name_shops


def get_engine(user, passwd, host, port, db):
    url = f"postgresql://{user}:{passwd}@{host}:{port}/{db}"
    if not database_exists(url):
        create_database(url)
    return create_engine(url, pool_size=50, echo=False)


def get_engine_from_settings():  # sourcery skip: raise-specific-error
    keys = ["user", "password", "host", "port", "database"]
    if any(key not in keys for key in settings.keys()):
        raise Exception("Файл настроек не правельный")
    return get_engine(settings["user"],
                      settings["password"],
                      settings["host"],
                      settings["port"],
                      settings["database"])


def get_session():
    engine = get_engine_from_settings()
    return sessionmaker(bind=engine)() # type: ignore


def connect_bd():
    return psycopg2.connect(
        database=settings["database"],
        user=settings["user"],
        password=settings["password"],
        host=settings["host"],
        port=settings["port"],
    )
    
def read_user(data_user):
    try:
        table_df = pd.read_sql(
            f"SELECT id FROM user_team WHERE id = {data_user['id']}", con=get_engine_from_settings())
        user_bool = len(table_df["id"])
        if user_bool >= 1:
            return "Приветствуем вас снова"
        elif user_bool == 0:
            created_user(data_user)
            return "Здравствуйте"
    except Exception as err:
        print(err)
        return "Произошла ошибка проверки личности"
    
def created_user(data_user):
    connection = connect_bd()
    connection.autocommit = True
    with connection.cursor() as cursor:
        cursor.execute("""INSERT INTO user_team(id, is_bot, first_name, username, language_code)
                          VALUES(%s, %s, %s, %s, %s)""", [data_user["id"], data_user["is_bot"], data_user["first_name"], data_user["username"],
                                                          data_user["language_code"]])
        
def get_sales_today():
    try:
        table_wb_cards = pd.read_sql(
            "SELECT * FROM price_comparison_wb_team WHERE percentage_difference <= -50",
            con=get_engine_from_settings(),
        )
        contact_user = pd.read_sql("SELECT contacts FROM user_team WHERE id IN(1323522063)", con=get_engine_from_settings()) #549779286
        full_list_contact = ["1323522063"] #549779286
        if contact_user["contacts"][0] != None:
            list_con = contact_user["contacts"][0].split(',')
            full_list_contact.extend(
                list_con[it_con] for it_con in range(len(list_con) - 1)
            )
        book = openpyxl.Workbook()
        del book['Sheet']
        sheet = book.create_sheet("Сводная по изменению продаж")
        sheet["A1"] = "Артикул"
        sheet["A1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 30
        sheet["B1"] = "Продано позавчера"
        sheet["B1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['B'].width = 18
        sheet["C1"] = "Продано вчера"
        sheet["C1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['C'].width = 16
        sheet["D1"] = "Разница продаж"
        sheet["D1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['D'].width = 16
        sheet["E1"] = "Процент разницы"
        sheet["E1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['E'].width = 16
        sheet["F1"] = "Магазин"
        sheet["F1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['F'].width = 16
        row = 2
        for insex, item in enumerate(list(table_wb_cards["articl"])):
            sheet[f"A{row}"] = item
            sheet[f"B{row}"] = list(table_wb_cards["sales_yesterday"])[insex]
            sheet[f"B{row}"].alignment = Alignment(horizontal='center')
            sheet[f"C{row}"] = list(table_wb_cards["sales_today"])[insex]
            sheet[f"C{row}"].alignment = Alignment(horizontal='center')
            sheet[f"D{row}"] = f"{list(table_wb_cards['sales_difference'])[insex]}  шт."
            sheet[f"D{row}"].alignment = Alignment(horizontal='center')
            sheet[f"E{row}"] = f"{list(table_wb_cards['percentage_difference'])[insex]}  %"
            sheet[f"E{row}"].alignment = Alignment(horizontal='center')
            sheet[f"F{row}"] = list(table_wb_cards["shops"])[insex]
            sheet[f"F{row}"].alignment = Alignment(horizontal='center')
            row += 1
        book.save("Изменение продаж.xlsx")
        book.close()
        return [list(table_wb_cards["articl"]), list(table_wb_cards["sales_yesterday"]), list(table_wb_cards["sales_today"]),
                list(table_wb_cards["sales_difference"]), list(table_wb_cards["percentage_difference"]),
                full_list_contact, list(table_wb_cards["shops"]), "Изменение продаж.xlsx"]
    except Exception as ex:
        logging.exception(ex)

def get_satat_wb():
    try:
        table_wb_cards = pd.read_sql(
            "SELECT * FROM no_sales_for_7_team WHERE stocks_user > 0 AND sales_user = 0",
            con=get_engine_from_settings(),
        )
        contact_user = pd.read_sql("SELECT contacts FROM user_team WHERE id IN(1323522063)", con=get_engine_from_settings())#549779286, 
        full_list_contact = ["1323522063"]# 549779286, 
        if contact_user["contacts"][0] != None:
            list_con = contact_user["contacts"][0].split(',')
            full_list_contact.extend(
                list_con[it_con] for it_con in range(len(list_con) - 1)
            )
        book = openpyxl.Workbook()
        del book['Sheet']
        sheet = book.create_sheet("Сводная не продаж за неделю")
        sheet["A1"] = "Артикул"
        sheet["A1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 30
        sheet["B1"] = "Название"
        sheet["B1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['B'].width = 18
        sheet["C1"] = "На складах"
        sheet["C1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['C'].width = 16
        sheet["D1"] = "Продано за неделю"
        sheet["D1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['D'].width = 20
        sheet["E1"] = "Ссылка"
        sheet["E1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['E'].width = 50
        sheet["F1"] = "Магазин"
        sheet["F1"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['F'].width = 16
        row = 2
        for insex, item in enumerate(list(table_wb_cards["supplierarticle"])):
            sheet[f"A{row}"] = item
            sheet[f"B{row}"] = list(table_wb_cards["subject"])[insex]
            sheet[f"C{row}"] = f"{list(table_wb_cards['stocks_user'])[insex]}  шт."
            sheet[f"C{row}"].alignment = Alignment(horizontal='center')
            sheet[f"D{row}"] = f"{list(table_wb_cards['sales_user'])[insex]}  шт."
            sheet[f"D{row}"].alignment = Alignment(horizontal='center')
            sheet[f"E{row}"] = list(table_wb_cards["link_site"])[insex]
            sheet[f"E{row}"].alignment = Alignment(horizontal='center')
            sheet[f"F{row}"] = list(table_wb_cards["shops"])[insex]
            sheet[f"F{row}"].alignment = Alignment(horizontal='center')
            row += 1
        book.save("Не проданные товары за неделю.xlsx")
        book.close()
        return [list(table_wb_cards["nmid"]), list(table_wb_cards["supplierarticle"]), list(table_wb_cards["subject"]),
                list(table_wb_cards["stocks_user"]), list(table_wb_cards["sales_user"]), list(table_wb_cards["link_site"]),
                full_list_contact, list(table_wb_cards["shops"]), "Не проданные товары за неделю.xlsx"]
    except Exception as ex:
        logging.exception(ex)
        
def created_xlsx_user_provider():  # sourcery skip: low-code-quality
    try:
        table_user = pd.read_sql("SELECT * FROM user_team WHERE id IN(1323522063)", con=get_engine_from_settings())#549779286, 
        full_list_contact = ['1323522063']#549779286, 
        if table_user["contacts"][0] != None:
            list_con = table_user["contacts"][0].split(',')
            full_list_contact.extend(
                list_con[it_con] for it_con in range(len(list_con) - 1)
            )
        return full_list_contact
    except Exception as ex:
        logging.exception(ex)

def create_comtact(id_user, contact):
    # sourcery skip: use-fstring-for-concatenation
    try:
        str_id_cont = str(contact)
        table_contact = pd.read_sql(
            f"SELECT contacts FROM user_team WHERE id = {id_user}", con=get_engine_from_settings())
        contact_str = table_contact["contacts"][0]
        connection = connect_bd()
        cursor = connection.cursor()
        sql_update_query = """Update user_team set contacts = %s where id = %s"""
        if contact_str is None:
            cursor.execute(sql_update_query, (str_id_cont + ',', id_user))
        else:
            list_contact = contact_str.split(",")
            if str_id_cont in list_contact:
                return f"Контакт {str_id_cont} у вас уже есть"
            cursor.execute(sql_update_query, (contact_str + str_id_cont + ',', id_user))
        connection.commit()
        return f"Контакт {str_id_cont} добавлен в ваши контакты"
    except Exception as ex:
        logging.exception(ex)
        
def get_contact(id):
    try:
        table_contact = pd.read_sql(
            f"SELECT contacts FROM user_team WHERE id = {id}", con=get_engine_from_settings())
        contact_str = table_contact["contacts"][0].split(",")
        list_first_name = []
        list_username = []
        list_id = []
        for item_con in range(len(contact_str) - 1):
            contact = pd.read_sql(
                f"SELECT * FROM user_team WHERE id = {contact_str[item_con]}", con=get_engine_from_settings())
            try:
                list_first_name.append(contact["first_name"][0])
                list_username.append(contact["username"][0])
            except Exception:
                list_first_name.append("Контакт не добавился в чат")
                list_username.append("Контакт не добавился в чат")
            list_id.append(contact_str[item_con])
        return [list_first_name, list_username, list_id]
    except Exception as ex:
        logging.exception(ex)
        
def delet_cont(id, contact):
    try:
        connection = connect_bd()
        cursor = connection.cursor()
        table_contact = pd.read_sql(
            f"SELECT contacts FROM user_team WHERE id = {id}", con=get_engine_from_settings())
        connt = contact + ','
        now_cont = table_contact["contacts"][0].replace(connt, '')
        sql_update_query = """Update user_team set contacts = %s where id = %s"""
        cursor.execute(sql_update_query, (now_cont, id))
        connection.commit()
        return f"Контакт {contact} удален"
    except Exception as ex:
        logging.exception(ex)
        return f"Контакт {contact} не удален. Проверьте список контактов, нажав кнопку 'Удалить контакты для рассылки'"

if __name__=="__main__":
    get_sales_today()
    